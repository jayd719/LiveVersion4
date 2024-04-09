from openpyxl import load_workbook
from openpyxl.comments import Comment
from LiveVersion4.settings import BASE_DIR
import os

AUTHOR ='JD'

def createExcelSheet(jobList,workCenter):
    wb = load_workbook(os.path.join(BASE_DIR,'files','static','shd','Scheduling Format.xlsx'))
    schedule = wb[wb.sheetnames[0]]
    startRow =2
    for workOrder in jobList:
        schedule[f'B{startRow}'] = workOrder.jobNumber.jobNumber
        schedule[f'B{startRow}'].comment =  Comment(text=str(workOrder.jobNumber.des), author=AUTHOR)
        schedule[f'C{startRow}'] = workOrder.jobNumber.qty
        schedule[f'E{startRow}'] = workOrder.workCenter
        schedule[f'F{startRow}'] = workOrder.stepNumber
        schedule[f'F{startRow}'].comment = Comment(text=str(workOrder.description), author=AUTHOR)
        schedule[f'U{startRow}'] = workOrder.jobNumber.dueDate
        schedule[f'V{startRow}'] = workOrder.jobNumber.customer
        schedule[f'W{startRow}'] = workOrder.jobNumber.TA
        schedule[f'X{startRow}'] = workOrder.jobNumber.notes1

        startRow+=1
    wb.save(os.path.join(BASE_DIR,'files','static','excel',f'{workCenter}.xlsx'))


